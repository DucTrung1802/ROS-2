from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import os


class DataRecoder(object):
    def __init__(self, pwm, pwm_frequency, sample_time):
        self.__initializeWorkBook()
        self.__preconfigure(pwm, pwm_frequency, sample_time)

    def __preconfigure(self, pwm, pwm_frequency, sample_time):
        self.__name = (
            "Motor_Data_"
            + str(pwm_frequency)
            + "Hz"
            + "_"
            + "Vary_PWM"
            + "_"
            + str(sample_time)
            + "s"
            + ".xlsx"
        )

    def __initializeWorkBook(self):
        self.__work_book = Workbook()
        for i in range(1, 12):
            self.__work_book.active.column_dimensions[get_column_letter(i)].width = 20
        # self.writeData(1, 1, "Time (s)")
        # self.writeData(1, 2, "Setpoint Motor 1 (RPM)")
        # self.writeData(1, 3, "Motor 1 (RPM)")
        # self.writeData(1, 4, "Ki Motor 1")

        # self.writeData(1, 6, "Setpoint Motor 2 (RPM)")
        # self.writeData(1, 7, "Motor 2 (RPM)")
        # self.writeData(1, 8, "Ki Motor 2")

        # self.writeData(1, 9, "Successful Receive")
        # self.writeData(1, 10, "Error Receive")
        # self.writeData(1, 11, "Rate of Successful (%)")

        self.writeData(1, 1, "Time (s)")
        self.writeData(1, 2, "Setpoint Motor 1 (RPM)")
        self.writeData(1, 3, "Motor 1 (RPM)")
        self.writeData(1, 4, "Kp Motor 1")
        self.writeData(1, 5, "Ki Motor 1")
        self.writeData(1, 6, "Kd Motor 1")

        self.writeData(1, 8, "Setpoint Motor 2 (RPM)")
        self.writeData(1, 9, "Motor 2 (RPM)")
        self.writeData(1, 10, "Kp Motor 2")
        self.writeData(1, 11, "Ki Motor 2")
        self.writeData(1, 12, "Kd Motor 2")

        # self.writeData(1, 9, "Successful Receive")
        # self.writeData(1, 10, "Error Receive")
        # self.writeData(1, 11, "Rate of Successful (%)")

    def saveWorkBook(self):
        # self.__folder_name = os.path.dirname(os.path.realpath(__file__))
        self.__save_path = os.path.join(Path.cwd(), self.__name)
        self.__work_book.save(filename=self.__save_path)
        print(self.__save_path)

    def writeData(self, row, col, data):
        self.__work_book.active.cell(row=row, column=col, value=data)


def main():
    i = 0
    workbook = DataRecoder("Motor_Data")
    workbook.configure(1023, 1000, 0.05)
    workbook.writeData(i + 5, 1, 10)
    workbook.saveWorkBook()


if __name__ == "__main__":
    main()
